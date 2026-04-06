from typing import Dict, Any
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.comments import Comment
from app.core.state import WorkflowState  # 关键：修复 NameError

def edit_and_save_excel_node(state: WorkflowState) -> Dict[str, Any]:
    print("--- 节点 4: 标记错误并保存 ---")
    
    results = state.get('llm_raw_response', {})
    input_file = state.get('file_path_v2')
    output_file = "data/output_marked.xlsx"
    
    if not input_file:
        print("错误：State 中未找到输入文件路径")
        return {}

    # 加载工作簿
    wb = load_workbook(input_file)
    ws = wb.active
    
    # 定义红色填充样式（用于错误单元格背景）
    red_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
    
    # 假设第一列是员工 ID (emp_id)
    id_column_index = 1 
    
    for row in range(2, ws.max_row + 1):
        # 获取当前行的 ID
        cell_obj = ws.cell(row=row, column=id_column_index)
        emp_id = str(cell_obj.value)
        
        # 检查该 ID 是否在 LLM 的错误结果中
        if emp_id in results and results[emp_id].get('is_error'):
            reason = results[emp_id].get('reason', '未知错误')
            
            # 1. 改变该单元格背景色（仅限 ID 单元格，不改整行）
            cell_obj.fill = red_fill
            
            # 2. 将错误原因写入单元格注释（不覆盖单元格原始内容）
            if not cell_obj.comment:
                cell_obj.comment = Comment(reason, "AI_System")
            else:
                cell_obj.comment.text += f"\n[新错误]: {reason}"
                
    # 3. 保存为新文件，不改变原始输入数据
    wb.save(output_file)
    print(f"处理完成，结果已保存至: {output_file}")
    
    return {"marked_file_path": output_file}