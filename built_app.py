import os

# 定义工程结构和内容
PROJECT_STRUCTURE = {
    # 配置文件
    "config/rules.txt": """1. 工资规范：所有员工的工资必须大于 0。
2. 部门限制：公司目前只设有 [技术部, 市场部, 财务部, 行政部]，若出现其他部门视为错误。
3. 日期逻辑：入职日期不能晚于 2026-12-31。
4. 变更备注：如果工资发生变更，涨幅不得超过 50%。""",

    ".env": "GOOGLE_API_KEY=你的_GEMINI_API_KEY_在这里",

    # 核心状态定义
    "app/core/state.py": """from typing import TypedDict, List, Dict, Any, Optional

class WorkflowState(TypedDict):
    file_path_v1: str
    file_path_v2: str
    diff_data: List[Dict[str, Any]]
    final_prompt: str
    llm_raw_response: Dict[str, Any]
    error_message: Optional[str]""",

    # 节点1: 数据提取
    "app/nodes/n1_data_extractor.py": """import pandas as pd
from typing import Dict, Any
from app.core.state import WorkflowState

def extract_diff_data_node(state: WorkflowState) -> Dict[str, Any]:
    print("--- 节点 1: 正在对比 Excel ---")
    df1 = pd.read_excel(state['file_path_v1'])
    df2 = pd.read_excel(state['file_path_v2'])
    
    # 简单对比：找出 V2 中不在 V1 的行，或者 ID 相同但内容不同的行
    # 这里为了演示，取 V2 全量数据作为待校验项
    diff_list = df2.to_dict(orient='records')
    return {"diff_data": diff_list}""",

    # 节点2: Prompt构建
    "app/nodes/n2_prompt_builder.py": """import json
from typing import Dict, Any
from app.core.state import WorkflowState

def build_prompt_node(state: WorkflowState) -> Dict[str, Any]:
    print("--- 节点 2: 拼接提示词 ---")
    with open('config/rules.txt', 'r', encoding='utf-8') as f:
        rules = f.read()
    
    data_str = json.dumps(state['diff_data'], ensure_ascii=False, indent=2)
    prompt = f"请根据规则校验数据。规则:\\n{rules}\\n\\n数据:\\n{data_str}\\n\\n请返回JSON:{{'ID':{{'is_error':bool, 'reason':str}}}}"
    return {"final_prompt": prompt}""",

    # 节点3: LLM调用
    "app/nodes/n3_llm_invoker.py": """import json
import os
from typing import Dict, Any
from langchain_google_genai import ChatGoogleGenerativeAI
from langchain_core.messages import HumanMessage
from app.core.state import WorkflowState

def invoke_gemini_node(state: WorkflowState) -> Dict[str, Any]:
    print("--- 节点 3: 调用 Gemini ---")
    llm = ChatGoogleGenerativeAI(model="gemini-1.5-flash", temperature=0)
    response = llm.invoke([HumanMessage(content=state['final_prompt'])])
    
    # 简单清洗并解析 JSON
    text = response.content.replace("```json", "").replace("```", "").strip()
    return {"llm_raw_response": json.loads(text)}""",

    # 节点4: Excel回写
    "app/nodes/n4_excel_editor.py": """from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from typing import Dict, Any
from app.core.state import WorkflowState

def edit_and_save_excel_node(state: WorkflowState) -> Dict[str, Any]:
    print("--- 节点 4: 标记错误并保存 ---")
    results = state['llm_raw_response']
    wb = load_workbook(state['file_path_v2'])
    ws = wb.active
    red_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
    
    for row in range(2, ws.max_row + 1):
        emp_id = str(ws.cell(row=row, column=1).value)
        if emp_id in results and results[emp_id].get('is_error'):
            for cell in ws[row]:
                cell.fill = red_fill
            ws.cell(row=row, column=ws.max_column).value = results[emp_id].get('reason')
            
    wb.save("data/output_marked.xlsx")
    return {}""",

    # 工作流编排
    "app/core/agent.py": """from langgraph.graph import StateGraph, END
from app.core.state import WorkflowState
from app.nodes.n1_data_extractor import extract_diff_data_node
from app.nodes.n2_prompt_builder import build_prompt_node
from app.nodes.n3_llm_invoker import invoke_gemini_node
from app.nodes.n4_excel_editor import edit_and_save_excel_node

def create_workflow():
    workflow = StateGraph(WorkflowState)
    workflow.add_node("n1", extract_diff_data_node)
    workflow.add_node("n2", build_prompt_node)
    workflow.add_node("n3", invoke_gemini_node)
    workflow.add_node("n4", edit_and_save_excel_node)
    
    workflow.set_entry_point("n1")
    workflow.add_edge("n1", "n2")
    workflow.add_edge("n2", "n3")
    workflow.add_edge("n3", "n4")
    workflow.add_edge("n4", END)
    return workflow.compile()""",

    # 主入口
    "main.py": """import os
import pandas as pd
from dotenv import load_dotenv
from app.core.agent import create_workflow

load_dotenv()

def prepare_test_excel():
    os.makedirs('data', exist_ok=True)
    df1 = pd.DataFrame({'ID': ['101'], '姓名': ['张三'], '工资': [5000]})
    df2 = pd.DataFrame({'ID': ['101', '102'], '姓名': ['张三', '李四'], '工资': [5200, 0]})
    df1.to_excel('data/v1.xlsx', index=False)
    df2.to_excel('data/v2.xlsx', index=False)

if __name__ == "__main__":
    prepare_test_excel()
    app = create_workflow()
    app.invoke({
        "file_path_v1": "data/v1.xlsx",
        "file_path_v2": "data/v2.xlsx",
        "diff_data": [], "final_prompt": "", "llm_raw_response": {}, "error_message": None
    })
    print("执行成功！请查看 data/output_marked.xlsx")"""
}

def build():
    for path, content in PROJECT_STRUCTURE.items():
        # 创建目录
        dir_name = os.path.dirname(path)
        if dir_name:
            os.makedirs(dir_name, exist_ok=True)
        
        # 写入文件
        with open(path, "w", encoding="utf-8") as f:
            f.write(content)
        print(f"已生成: {path}")

    # 创建必要的 __init__.py
    for d in ["app", "app/core", "app/nodes"]:
        with open(os.path.join(d, "__init__.py"), "w") as f:
            pass

if __name__ == "__main__":
    build()